"""Excel export: a stakeholder-facing summary workbook built from the DB.

Sheets: Fleet KPIs, Utilization by Asset, Downtime & MTBF, Fuel Efficiency,
Maintenance Due, Open Flags. Overdue / low-availability / low-utilization rows
are shaded red.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.engine import Engine

from .db import REPO_ROOT
from .metrics import compute_all, load_frames

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
BAD_FILL = PatternFill("solid", fgColor="F8D7DA")


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 42)
    return ws


def _shade(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        ws.cell(row=row_idx, column=c).fill = BAD_FILL


def build_report(engine: Engine, out_dir: str | Path | None = None, as_of: date | None = None) -> Path:
    frames = load_frames(engine)
    metrics = compute_all(frames, as_of=as_of)
    tags = sorted(metrics["maintenance"].keys())
    asset_meta = {a["assetTag"]: a for a in frames["assets"]}

    wb = Workbook()
    wb.remove(wb.active)

    k = metrics["fleet_kpis"]
    _sheet(wb, "Fleet KPIs", ["Metric", "Value"], [
        ["As of", metrics["as_of"]],
        ["Assets", k["assets"]],
        ["Avg utilization %", k["avg_utilization_pct"]],
        ["Avg availability %", k["avg_availability_pct"]],
        ["Avg fuel cost / engine-hr", k["avg_fuel_cost_per_engine_hour"]],
        ["Total unplanned downtime (h, 30d)", k["total_unplanned_downtime_hours"]],
        ["Assets overdue for service", k["assets_overdue_service"]],
    ])

    ws = _sheet(
        wb, "Utilization by Asset",
        ["Asset", "Type", "Days logged", "Engine hrs (30d)", "Avg eng-hrs/day", "Avg utilization %"],
        [[
            t, asset_meta[t]["type"],
            metrics["utilization"][t]["days_logged"],
            metrics["utilization"][t]["total_engine_hours_30d"],
            metrics["utilization"][t]["avg_engine_hours_per_day"],
            metrics["utilization"][t]["avg_utilization_pct_30d"],
        ] for t in tags],
    )
    for i, t in enumerate(tags, start=2):
        v = metrics["utilization"][t]["avg_utilization_pct_30d"]
        if v is not None and v < 45:
            _shade(ws, i, 6)

    ws = _sheet(
        wb, "Downtime & MTBF",
        ["Asset", "Unplanned events", "Unplanned hrs (30d)", "Total downtime hrs", "MTBF hrs", "MTTR hrs", "Availability %"],
        [[
            t,
            metrics["downtime"][t]["unplanned_events_30d"],
            metrics["downtime"][t]["unplanned_downtime_hours_30d"],
            metrics["downtime"][t]["total_downtime_hours_30d"],
            metrics["downtime"][t]["mtbf_hours"],
            metrics["downtime"][t]["mttr_hours"],
            metrics["downtime"][t]["availability_pct_30d"],
        ] for t in tags],
    )
    for i, t in enumerate(tags, start=2):
        if (metrics["downtime"][t]["availability_pct_30d"] or 100) < 80:
            _shade(ws, i, 7)

    _sheet(
        wb, "Fuel Efficiency",
        ["Asset", "Type", "Litres (30d)", "Fuel $ (30d)", "Engine hrs (30d)", "L / engine-hr", "$ / engine-hr"],
        [[
            t, asset_meta[t]["type"],
            metrics["fuel"][t]["litres_30d"],
            metrics["fuel"][t]["fuel_cost_30d"],
            metrics["fuel"][t]["engine_hours_30d"],
            metrics["fuel"][t]["litres_per_engine_hour"],
            metrics["fuel"][t]["cost_per_engine_hour"],
        ] for t in tags],
    )

    ws = _sheet(
        wb, "Maintenance Due",
        ["Asset", "Service interval hrs", "Hrs since service", "Hrs to next service", "Overdue"],
        [[
            t,
            metrics["maintenance"][t]["service_interval_hours"],
            metrics["maintenance"][t]["hours_since_service"],
            metrics["maintenance"][t]["hours_to_next_service"],
            "YES" if metrics["maintenance"][t]["service_overdue"] else "",
        ] for t in tags],
    )
    for i, t in enumerate(tags, start=2):
        if metrics["maintenance"][t]["service_overdue"]:
            _shade(ws, i, 5)

    with engine.connect() as cx:
        flags = [dict(r._mapping) for r in cx.execute(text("""
            SELECT a."assetTag", f."kind", f."status", f."observedValue", f."thresholdValue",
                   f."externalTicketNumber", f."detail", f."createdAt"
            FROM "MaintenanceFlag" f JOIN "Asset" a ON a."id" = f."assetId"
            WHERE f."status" <> 'RESOLVED'
            ORDER BY f."createdAt" DESC
        """))]
    _sheet(
        wb, "Open Flags",
        ["Asset", "Kind", "Status", "Observed", "Threshold", "Ticket", "Detail", "Raised"],
        [[
            r["assetTag"], r["kind"], r["status"], r["observedValue"], r["thresholdValue"],
            r["externalTicketNumber"] or "", r["detail"],
            r["createdAt"].isoformat(sep=" ", timespec="minutes") if r["createdAt"] else "",
        ] for r in flags] or [["—", "", "", "", "", "", "no open flags", ""]],
    )

    out_dir = Path(out_dir) if out_dir else REPO_ROOT / "data" / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"fleet-summary-{metrics['as_of']}.xlsx"
    wb.save(path)
    return path
