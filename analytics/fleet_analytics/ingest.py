"""Excel ingest: read a raw fleet workbook, validate it, and upsert into Postgres.

Sheets expected (header row 1):
  Assets           assetTag,name,type,site,model,commissionedAt,scheduledHoursPerDay,
                   serviceIntervalHours,lastServiceHours,currentEngineHours,status
  UtilizationLogs  assetTag,date,engineHours,idleHours,distanceKm,payloadTonnes
  FuelLogs         assetTag,date,litres,cost,engineHoursAtFill
  DowntimeEvents   assetTag,startAt,endAt,category,reason,hours

Semantics: Assets / UtilizationLogs / FuelLogs are upserted (idempotent by natural
key). DowntimeEvents for the assets in the file are replaced.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Engine

ASSET_TYPES = {"HAUL_TRUCK", "EXCAVATOR", "DOZER", "LOADER", "GRADER", "WATER_TRUCK"}
ASSET_STATUSES = {"ACTIVE", "DOWN", "MAINTENANCE", "RETIRED"}
DOWNTIME_CATEGORIES = {"UNPLANNED", "PLANNED", "STANDBY"}

REQUIRED_SHEETS = {
    "Assets": [
        "assetTag", "name", "type", "site", "model", "commissionedAt",
        "scheduledHoursPerDay", "serviceIntervalHours", "lastServiceHours",
        "currentEngineHours", "status",
    ],
    "UtilizationLogs": ["assetTag", "date", "engineHours", "idleHours", "distanceKm", "payloadTonnes"],
    "FuelLogs": ["assetTag", "date", "litres", "cost", "engineHoursAtFill"],
    "DowntimeEvents": ["assetTag", "startAt", "endAt", "category", "reason", "hours"],
}


class ValidationError(Exception):
    def __init__(self, problems: list[str]):
        self.problems = problems
        super().__init__(f"{len(problems)} validation problem(s):\n  - " + "\n  - ".join(problems))


# ── reading ─────────────────────────────────────────────────────────────────

def _read_sheet(wb, name: str) -> list[dict[str, Any]]:
    if name not in wb.sheetnames:
        raise ValidationError([f"missing required sheet '{name}'"])
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


def read_workbook(path: str | Path) -> dict[str, list[dict]]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return {name: _read_sheet(wb, name) for name in REQUIRED_SHEETS}
    finally:
        wb.close()


# ── coercion helpers ───────────────────────────────────────────────────────

def _num(v):
    if v is None or v == "":
        return None
    return float(v)


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date.fromisoformat(str(v)[:10])


def _to_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return datetime.fromisoformat(str(v).replace("Z", "+00:00").replace(" ", "T")[:19])


# ── validation ─────────────────────────────────────────────────────────────

def validate(data: dict[str, list[dict]]) -> None:
    problems: list[str] = []

    for sheet, cols in REQUIRED_SHEETS.items():
        rows = data.get(sheet, [])
        if rows:
            missing = [c for c in cols if c not in rows[0]]
            if missing:
                problems.append(f"{sheet}: missing column(s) {missing}")

    assets = data.get("Assets", [])
    if not assets:
        problems.append("Assets: no rows")
    tags: set[str] = set()
    for i, a in enumerate(assets, start=2):
        tag = (a.get("assetTag") or "").strip()
        loc = f"Assets row {i}"
        if not tag:
            problems.append(f"{loc}: blank assetTag")
            continue
        if tag in tags:
            problems.append(f"{loc}: duplicate assetTag '{tag}'")
        tags.add(tag)
        if a.get("type") not in ASSET_TYPES:
            problems.append(f"{loc}: bad type {a.get('type')!r}")
        if (a.get("status") or "ACTIVE") not in ASSET_STATUSES:
            problems.append(f"{loc}: bad status {a.get('status')!r}")
        try:
            sh = _num(a.get("scheduledHoursPerDay"))
            if sh is None or not (0 < sh <= 24):
                problems.append(f"{loc}: scheduledHoursPerDay must be in (0, 24]")
            for f in ("serviceIntervalHours", "lastServiceHours", "currentEngineHours"):
                if (_num(a.get(f)) or 0) < 0:
                    problems.append(f"{loc}: {f} must be >= 0")
            _to_dt(a.get("commissionedAt"))
        except (ValueError, TypeError) as e:
            problems.append(f"{loc}: {e}")

    def check_logs(sheet, checks):
        for i, row in enumerate(data.get(sheet, []), start=2):
            loc = f"{sheet} row {i}"
            t = (row.get("assetTag") or "").strip()
            if t not in tags:
                problems.append(f"{loc}: unknown assetTag '{t}'")
                continue
            try:
                checks(row, loc)
            except (ValueError, TypeError) as e:
                problems.append(f"{loc}: {e}")

    def util_checks(row, loc):
        _to_date(row.get("date"))
        eh = _num(row.get("engineHours"))
        if eh is None or not (0 <= eh <= 24):
            problems.append(f"{loc}: engineHours must be in [0, 24]")
        if (_num(row.get("idleHours")) or 0) < 0:
            problems.append(f"{loc}: idleHours must be >= 0")

    def fuel_checks(row, loc):
        _to_date(row.get("date"))
        litres = _num(row.get("litres"))
        if litres is None or litres <= 0:
            problems.append(f"{loc}: litres must be > 0")
        if (_num(row.get("cost")) or 0) < 0:
            problems.append(f"{loc}: cost must be >= 0")

    def downtime_checks(row, loc):
        _to_dt(row.get("startAt"))
        if row.get("endAt") not in (None, ""):
            _to_dt(row.get("endAt"))
        if row.get("category") not in DOWNTIME_CATEGORIES:
            problems.append(f"{loc}: bad category {row.get('category')!r}")
        if (_num(row.get("hours")) or 0) < 0:
            problems.append(f"{loc}: hours must be >= 0")

    check_logs("UtilizationLogs", util_checks)
    check_logs("FuelLogs", fuel_checks)
    check_logs("DowntimeEvents", downtime_checks)

    if problems:
        raise ValidationError(problems)


# ── upsert ─────────────────────────────────────────────────────────────────

_ASSET_UPSERT = text("""
INSERT INTO "Asset" ("id","assetTag","name","type","site","model","commissionedAt",
  "scheduledHoursPerDay","serviceIntervalHours","lastServiceHours","currentEngineHours","status","updatedAt")
VALUES (:id,:assetTag,:name,CAST(:type AS "AssetType"),:site,:model,:commissionedAt,
  :scheduledHoursPerDay,:serviceIntervalHours,:lastServiceHours,:currentEngineHours,CAST(:status AS "AssetStatus"),now())
ON CONFLICT ("assetTag") DO UPDATE SET
  "name"=EXCLUDED."name","type"=EXCLUDED."type","site"=EXCLUDED."site","model"=EXCLUDED."model",
  "commissionedAt"=EXCLUDED."commissionedAt","scheduledHoursPerDay"=EXCLUDED."scheduledHoursPerDay",
  "serviceIntervalHours"=EXCLUDED."serviceIntervalHours","lastServiceHours"=EXCLUDED."lastServiceHours",
  "currentEngineHours"=EXCLUDED."currentEngineHours","status"=EXCLUDED."status","updatedAt"=now()
""")

_UTIL_UPSERT = text("""
INSERT INTO "UtilizationLog" ("id","assetId","date","engineHours","idleHours","distanceKm","payloadTonnes")
VALUES (:id,:assetId,:date,:engineHours,:idleHours,:distanceKm,:payloadTonnes)
ON CONFLICT ("assetId","date") DO UPDATE SET
  "engineHours"=EXCLUDED."engineHours","idleHours"=EXCLUDED."idleHours",
  "distanceKm"=EXCLUDED."distanceKm","payloadTonnes"=EXCLUDED."payloadTonnes"
""")

_FUEL_UPSERT = text("""
INSERT INTO "FuelLog" ("id","assetId","date","litres","cost","engineHoursAtFill")
VALUES (:id,:assetId,:date,:litres,:cost,:engineHoursAtFill)
ON CONFLICT ("assetId","date") DO UPDATE SET
  "litres"=EXCLUDED."litres","cost"=EXCLUDED."cost","engineHoursAtFill"=EXCLUDED."engineHoursAtFill"
""")

_DOWNTIME_INSERT = text("""
INSERT INTO "DowntimeEvent" ("id","assetId","startAt","endAt","category","reason","hours")
VALUES (:id,:assetId,:startAt,:endAt,CAST(:category AS "DowntimeCategory"),:reason,:hours)
""")


def _rid(prefix: str) -> str:
    return f"{prefix}_{uuid4().hex}"


def ingest_file(engine: Engine, path: str | Path) -> dict[str, int]:
    data = read_workbook(path)
    validate(data)

    counts = {"assets": 0, "utilization": 0, "fuel": 0, "downtime": 0}

    with engine.begin() as cx:
        for a in data["Assets"]:
            cx.execute(_ASSET_UPSERT, {
                "id": _rid("ing"),
                "assetTag": a["assetTag"].strip(),
                "name": a["name"],
                "type": a["type"],
                "site": a["site"],
                "model": a["model"],
                "commissionedAt": _to_dt(a["commissionedAt"]),
                "scheduledHoursPerDay": _num(a["scheduledHoursPerDay"]),
                "serviceIntervalHours": _num(a["serviceIntervalHours"]),
                "lastServiceHours": _num(a["lastServiceHours"]),
                "currentEngineHours": _num(a["currentEngineHours"]),
                "status": (a.get("status") or "ACTIVE"),
            })
            counts["assets"] += 1

        id_by_tag = {
            row.assetTag: row.id
            for row in cx.execute(text('SELECT "id","assetTag" FROM "Asset"'))
        }

        util_params = [{
            "id": _rid("ing"), "assetId": id_by_tag[r["assetTag"].strip()],
            "date": _to_date(r["date"]), "engineHours": _num(r["engineHours"]),
            "idleHours": _num(r.get("idleHours")) or 0.0,
            "distanceKm": _num(r.get("distanceKm")) or 0.0,
            "payloadTonnes": _num(r.get("payloadTonnes")),
        } for r in data["UtilizationLogs"]]
        if util_params:
            cx.execute(_UTIL_UPSERT, util_params)
        counts["utilization"] = len(util_params)

        fuel_params = [{
            "id": _rid("ing"), "assetId": id_by_tag[r["assetTag"].strip()],
            "date": _to_date(r["date"]), "litres": _num(r["litres"]),
            "cost": _num(r.get("cost")) or 0.0, "engineHoursAtFill": _num(r.get("engineHoursAtFill")) or 0.0,
        } for r in data["FuelLogs"]]
        if fuel_params:
            cx.execute(_FUEL_UPSERT, fuel_params)
        counts["fuel"] = len(fuel_params)

        dt_rows = data["DowntimeEvents"]
        affected = sorted({id_by_tag[r["assetTag"].strip()] for r in dt_rows})
        if affected:
            cx.execute(
                text('DELETE FROM "DowntimeEvent" WHERE "assetId" = ANY(:ids)'),
                {"ids": affected},
            )
        dt_params = [{
            "id": _rid("ing"), "assetId": id_by_tag[r["assetTag"].strip()],
            "startAt": _to_dt(r["startAt"]),
            "endAt": _to_dt(r["endAt"]) if r.get("endAt") not in (None, "") else None,
            "category": r["category"], "reason": r.get("reason") or "",
            "hours": _num(r.get("hours")) or 0.0,
        } for r in dt_rows]
        if dt_params:
            cx.execute(_DOWNTIME_INSERT, dt_params)
        counts["downtime"] = len(dt_params)

    return counts
